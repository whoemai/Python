from azure.identity import DefaultAzureCredential
from azure.mgmt.compute import ComputeManagementClient
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import concurrent.futures

# Lista de IDs das subscriptions
sub1 = "xxx"
sub2 = "xxx"
sub3 = "xxx"
sub4 = "xxx"
subscriptions  = [sub1, sub2, sub3 ,sub4]

# Configuração das credenciais da Azure
credential = DefaultAzureCredential()

# Inicialização da planilha do Excel
workbook = Workbook()
sheet = workbook.active
sheet.append(
    [
        "Hostname",
        "Subscription",
        "Resource Group",
        "Tags",
        "Status",
        "Localizacao",
        "VM Size",
        "Zona",
    ]
)

def status_servidor(subscription_id, resource_group_name, servidor):
    client = ComputeManagementClient(
        credential      = credential,
        subscription_id = subscription_id,
    )

    response = client.virtual_machines.instance_view(
        resource_group_name = resource_group_name,
        vm_name             = servidor,
    )

    statuses = response.statuses

    if statuses:
        primeiro_status = statuses[0]
        status_servidor = primeiro_status.display_status
        return status_servidor

    return "N/A"

def process_vm(subscription_id, resource_group, servidor):
    status = status_servidor(subscription_id, resource_group, servidor.name)

    if servidor.tags:
        formatted_tags = ", ".join(
            [f"{key}: {value}" for key, value in servidor.tags.items()]
        )
    else:
        formatted_tags = "N/A"

    formatted_zones = ", ".join(servidor.zones) if servidor.zones else "N/A"
    return [
        servidor.name,
        subscription_id,
        resource_group,
        formatted_tags,
        status,
        servidor.location,
        servidor.hardware_profile.vm_size,
        formatted_zones,
    ]

with concurrent.futures.ThreadPoolExecutor() as executor:
    futures = []
    for subscription_id in subscriptions:
        compute_client = ComputeManagementClient(credential, subscription_id)
        servidores = compute_client.virtual_machines.list_all()
        for servidor in servidores:
            resource_id_parts = servidor.id.split("/")
            resource_group_index = resource_id_parts.index("resourceGroups") + 1
            resource_group = resource_id_parts[resource_group_index]

            if resource_group == "xxx":
                continue  # Pula para a próxima iteração ignorando o rg desejado

            subscription_index = resource_id_parts.index("subscriptions") + 1
            subscription       = resource_id_parts[subscription_index]

            futures.append(
                executor.submit(process_vm, subscription, resource_group, servidor)
            )

    for future in concurrent.futures.as_completed(futures):
        sheet.append(future.result())

# Formatação da largura das colunas
for column in sheet.columns:
    max_length     = max(len(str(cell.value)) for cell in column)
    adjusted_width = max_length + 2
    column_letter  = get_column_letter(column[0].column)
    sheet.column_dimensions[column_letter].width = adjusted_width

# Salvar a planilha do Excel
workbook.save("informacoes_servidores.xlsx")
print("Arquivo salvo com sucesso")